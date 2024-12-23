import requests
from datetime import datetime
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
import openpyxl
from openpyxl import load_workbook
import time

API_KEY = 'b0825721-aa7d-4ea2-ac8e-84a8a2c65eaf'
API_URL = 'https://pro-api.coinmarketcap.com/v1/cryptocurrency/listings/latest'
PDF_FILE = 'PythonAssignment/CryptoAnalysisReport.pdf'
EXCEL_FILE = 'PythonAssignment/CryptoLiveData.xlsx'
def fetch_crypto_data():
    try:
        response = requests.get(
            API_URL,
            headers={'X-CMC_PRO_API_KEY': API_KEY},
            params={'start': '1', 'limit': '50', 'convert': 'USD'}
        )
        response.raise_for_status()
        return response.json()['data']
    except requests.RequestException as e:
        print(f"Error fetching data: {e}")
        return []

def analyze_data(data):
    analysis = {}
    if not data:
        return analysis

    top_by_market_cap = sorted(data, key=lambda x: x['quote']['USD']['market_cap'], reverse=True)[:5]
    analysis['top_market_cap'] = top_by_market_cap

    most_volatile = sorted(data, key=lambda x: abs(x['quote']['USD']['percent_change_24h']), reverse=True)[:5]
    analysis['most_volatile'] = most_volatile

    total_market_cap = sum(crypto['quote']['USD']['market_cap'] for crypto in data)
    analysis['total_market_cap'] = total_market_cap

    analysis['summary'] = [
        f"Total market capitalization of top 50 cryptocurrencies: ${total_market_cap:,.2f}.",
        "Bitcoin remains dominant in market cap and trading volume.",
        "Altcoins show significant volatility, offering trading opportunities.",
        "Stablecoins like Tether ensure liquidity in the crypto market."
    ]
    return analysis

def generate_pdf_report(analysis):
    if not analysis:
        print("No data to generate the report.")
        return

    c = canvas.Canvas(PDF_FILE, pagesize=letter)
    width, height = letter

    c.setFont("Helvetica-Bold", 16)
    c.drawString(50, height - 50, "Cryptocurrency Market Analysis Report")
    c.setFont("Helvetica", 12)
    c.drawString(50, height - 70, f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")

    y = height - 110

    c.setFont("Helvetica-Bold", 14)
    c.drawString(50, y, "1. Top 5 Cryptocurrencies by Market Capitalization")
    y -= 20
    c.setFont("Helvetica", 12)
    for crypto in analysis['top_market_cap']:
        c.drawString(60, y, f"- {crypto['name']} ({crypto['symbol']}): ${crypto['quote']['USD']['market_cap']:.2f}")
        y -= 20

    c.setFont("Helvetica-Bold", 14)
    c.drawString(50, y, "2. Most Volatile Cryptocurrencies (24h Change %)")
    y -= 20
    c.setFont("Helvetica", 12)
    for crypto in analysis['most_volatile']:
        c.drawString(60, y, f"- {crypto['name']} ({crypto['symbol']}): {crypto['quote']['USD']['percent_change_24h']:.2f}%")
        y -= 20

    c.setFont("Helvetica-Bold", 14)
    c.drawString(50, y, "3. Summary of Insights")
    y -= 20
    c.setFont("Helvetica", 12)
    for point in analysis['summary']:
        c.drawString(60, y, f"- {point}")
        y -= 20
    c.save()


def save_to_excel(data):
    try:
        try:
            workbook = load_workbook(EXCEL_FILE)
            sheet = workbook.active
        except FileNotFoundError:
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = 'Crypto Data'
            headers = ['Name', 'Symbol', 'Price (USD)', 'Market Cap (USD)', '24h Volume (USD)', '24h Change (%)']
            sheet.append(headers)
        for crypto in data:
            sheet.append([
                crypto['name'],
                crypto['symbol'],
                crypto['quote']['USD']['price'],
                crypto['quote']['USD']['market_cap'],
                crypto['quote']['USD']['volume_24h'],
                crypto['quote']['USD']['percent_change_24h']
            ])
        workbook.save(EXCEL_FILE)
    except Exception as e:
        print(f"Error saving to Excel: {e}")


def main():

    while True:
        data = fetch_crypto_data()
        if data:
            save_to_excel(data)
            analysis = analyze_data(data)
            generate_pdf_report(analysis)
            print("Done at",datetime.now().date(),datetime.now().time())
        else:
            print("Failed to fetch data. Trying again in 5 minutes.")
        time.sleep(300) 

if __name__ == "__main__":
    main()
