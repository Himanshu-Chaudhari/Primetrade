import requests
import openpyxl
from openpyxl import load_workbook
import time

API_KEY = 'b0825721-aa7d-4ea2-ac8e-84a8a2c65eaf'
API_URL = 'https://pro-api.coinmarketcap.com/v1/cryptocurrency/listings/latest'
EXCEL_FILE = 'PythonAssignment/CryptoLiveData.xlsx'

def fetch_crypto_data():
    try:
        response = requests.get(
            API_URL,
            headers={'X-CMC_PRO_API_KEY': API_KEY},
            params={'start': '1', 'limit': '50', 'convert': 'USD'}
        )
        response.raise_for_status() 
        return response.json()['data']
    except requests.RequestException as e:
        print(f"Error fetching data: {e}")
        return []

def save_to_excel(data):
    try:
        try:
            workbook = load_workbook(EXCEL_FILE)
            sheet = workbook.active
        except FileNotFoundError:
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = 'Crypto Data'
            headers = ['Name', 'Symbol', 'Price (USD)', 'Market Cap (USD)', '24h Volume (USD)', '24h Change (%)']
            sheet.append(headers)
        for crypto in data:
            sheet.append([
                crypto['name'],
                crypto['symbol'],
                crypto['quote']['USD']['price'],
                crypto['quote']['USD']['market_cap'],
                crypto['quote']['USD']['volume_24h'],
                crypto['quote']['USD']['percent_change_24h']
            ])
        workbook.save(EXCEL_FILE)
        print(f"Data saved to {EXCEL_FILE}")
    except Exception as e:
        print(f"Error saving to Excel: {e}")

def main():
    while True:
        data = fetch_crypto_data()
        if data:
            save_to_excel(data)
        else:
            print("Failed to fetch data. Trying again in 5 minutes.")
        time.sleep(300) 

if __name__ == "__main__":
    main()
